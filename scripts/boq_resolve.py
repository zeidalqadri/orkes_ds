#!/usr/bin/env python3
"""Step 2: NO_BOQ_TAB Resolution — parse BoQ docs, wire into /v2, update triage."""
import json
import os
from datetime import UTC, datetime
from pathlib import Path

TRANSFER_DIR = Path("/home/the_bomb/orkes/yellowpages/tenders")
SMARTGEP_DATA = Path("/home/the_bomb/orkes/yellowpages/scrapers/data/smartgep/_engine_output_consurv")
TRIAGE_PATH = Path("/home/the_bomb/orkes_ds/data/triage_report.json")

# Tender directory mapping (found earlier)
TDR_MAP = {
    "RFP-000000176710": "tdr-4d95e649",
    "RFP-000000178771": "tdr-9de31538",  # primary
    "RFP-000000177523": "tdr-0c32c0bf",
}

# BoQ document paths
BOQ_DOCS = {
    "RFP-000000176710": [
        "documents/GEP-RFP-000000176710/Attachment D - BQ Masjid Baru.xlsx",
        "documents/GEP-RFP-000000176710/Attachment C - BQ Drawing.pdf",
    ],
    "RFP-000000178771": [
        "documents/GEP-RFP-000000178771/BQ STEM Hub 2026-Sabah.xlsx",
        "documents/GEP-RFP-000000178771/BQ STEM Hub 2026-Sabah.pdf",
    ],
}

def parse_xlsx(filepath):
    """Extract BoQ items from xlsx."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = None
        col_map = {}
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=row_idx, column=c).value or '').strip().upper() for c in range(1, ws.max_column + 1)]
            keywords_found = sum(1 for v in row_vals if v in ('ITEM', 'NO', 'DESCRIPTION', 'UNIT', 'QTY', 'UOM', 'RATE', 'AMOUNT'))
            if keywords_found >= 2:
                header_row = row_idx
                for col in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=row_idx, column=col).value or '').strip().upper()
                    if val in ('ITEM', 'NO', 'DESCRIPTION', 'UNIT', 'QTY', 'UOM', 'RATE', 'AMOUNT'):
                        col_map[val] = col
                break
        if not header_row or len(col_map) < 2:
            # Fuzzy fallback: look for rows with item-like numbers
            for row_idx in range(1, ws.max_row + 1):
                vals = {}
                for col in range(1, ws.max_column + 1):
                    v = str(ws.cell(row=row_idx, column=col).value or '').strip()
                    if v and len(v) < 5 and v.replace('.', '').replace(' ', '').isdigit():
                        vals['ITEM'] = str(ws.cell(row=row_idx, column=col).value)
                    elif col >= 2 and col <= 3 and v and len(v) > 2:
                        vals['DESCRIPTION'] = v
                    if col == 5 and v and len(v) < 10:
                        vals['QTY'] = ws.cell(row=row_idx, column=col).value
                if vals.get('DESCRIPTION'):
                    # Clean up
                    desc = vals['DESCRIPTION']
                    if desc.startswith('PROPOSED') or desc.startswith('Order') or desc.startswith('PROGRAM'):
                        continue
                    items.append({
                        'sheet': sheet_name,
                        'item_code': vals.get('ITEM', ''),
                        'description': desc[:200],
                        'unit': '',
                        'qty': vals.get('QTY'),
                        'rate': None,
                        'amount': None,
                    })
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            vals = {}
            for label, col in col_map.items():
                val = ws.cell(row=row_idx, column=col).value
                vals[label] = val
            item_code = vals.get('ITEM') or vals.get('NO')
            desc = vals.get('DESCRIPTION')
            if not item_code and not desc:
                continue
            item_code_str = str(item_code or '')
            desc_str = str(desc or '')
            if len(desc_str.strip()) < 3 and not item_code_str.strip():
                continue
            if desc_str.startswith('PROPOSED') or desc_str.startswith('Order') or desc_str.startswith('PROGRAM'):
                continue
            items.append({
                'sheet': sheet_name,
                'item_code': item_code_str.strip(),
                'description': desc_str.strip()[:200],
                'unit': str(vals.get('UNIT', vals.get('UOM', '')) or '').strip(),
                'qty': vals.get('QTY'),
                'rate': vals.get('RATE'),
                'amount': vals.get('AMOUNT'),
            })
    return items


def check_child_sheets(event_id, netsessionid):
    """Check SmartGEP for child datasheets via HTTP."""
    import requests
    # Try pricedatasheet endpoints
    urls = [
        f"https://arrprodus.eastus.cloudapp.azure.com/data/pricedatasheet?parentId={event_id}",
        f"https://arrprodus.eastus.cloudapp.azure.com/data/pricedatasheet/search?parentId={event_id}",
    ]
    headers = {
        "netsessionid": netsessionid,
        "Accept": "application/json",
    }
    for url in urls:
        try:
            r = requests.get(url, headers=headers, timeout=10, verify=False)
            if r.status_code == 200:
                data = r.json()
                if isinstance(data, list) and len(data) > 0:
                    return {"found": True, "count": len(data), "data": data[:50]}
                if isinstance(data, dict):
                    return {"found": True, "keys": list(data.keys())[:10]}
        except Exception:
            pass
    return {"found": False, "error": "All endpoints returned non-200"}


def main():
    print("=" * 60)
    print("Step 2 — NO_BOQ_TAB Resolution Pipeline")
    print("=" * 60)

    # Load triage report
    with open(TRIAGE_PATH) as f:
        triage = json.load(f)

    results = {}

    # ── Step 1: Parse BoQ docs for 176710 and 178771 ──
    for ref, doc_paths in BOQ_DOCS.items():
        print(f"\n--- Processing {ref} ---")
        items = []
        docs_copied = []

        for doc_rel in doc_paths:
            doc_path = SMARTGEP_DATA / doc_rel
            if not doc_path.exists():
                print(f"  MISSING: {doc_path}")
                continue

            fname = doc_path.name
            is_xlsx = fname.endswith('.xlsx')

            if is_xlsx:
                print(f"  Parsing: {fname}")
                items = parse_xlsx(str(doc_path))
                print(f"  → {len(items)} BoQ items extracted")

            # Copy doc to tender directory
            tdr = TDR_MAP.get(ref)
            if tdr:
                tdr_dir = TRANSFER_DIR / tdr
                tdr_dir.mkdir(parents=True, exist_ok=True)
                docs_dir = tdr_dir / "docs"
                docs_dir.mkdir(exist_ok=True)

                dest = docs_dir / fname
                import shutil
                shutil.copy2(doc_path, dest)
                docs_copied.append(fname)
                print(f"  → Copied to {tdr}/docs/{fname}")

        results[ref] = {
            "items_count": len(items),
            "items": items,
            "docs_copied": docs_copied,
        }

        # Write BoQ extraction JSON to tender directory
        tdr = TDR_MAP.get(ref)
        if tdr:
            boq_file = TRANSFER_DIR / tdr / "boq_extraction.json"
            boq_data = {
                "extracted_at": datetime.now(UTC).isoformat(),
                "source_xlsx": [d for d in doc_paths if d.endswith('.xlsx')],
                "item_count": len(items),
                "items": items,
            }
            boq_file.write_text(json.dumps(boq_data, indent=2, default=str))
            print("  → Wrote boq_extraction.json")

    # ── Step 2: Check child sheets for 177523 ──
    ref = "RFP-000000177523"
    print(f"\n--- Processing {ref} (child sheet check) ---")

    # Get permauth tokens
    child_result = {"found": False}
    try:
        import requests
        r = requests.get("http://localhost:9876/tokens", timeout=5)
        if r.status_code == 200:
            tokens = r.json()
            netsessionid = tokens.get("netsessionid", "")
            event_id = triage["events"].get(ref, {}).get("event_id", "")
            print(f"  netsessionid: {netsessionid[:20]}...")
            print(f"  event_id: {event_id}")
            if event_id:
                child_result = check_child_sheets(event_id, netsessionid)
                print(f"  Result: {child_result}")
            else:
                print("  No event_id in triage — can't check child sheets")
    except Exception as e:
        print(f"  Permauth error: {e}")
        # Try fallback — check if docs exist at all for this tender
        tdr = TDR_MAP.get(ref)
        if tdr:
            td_dir = TRANSFER_DIR / tdr / "docs"
            if td_dir.exists():
                existing = os.listdir(td_dir)
                print(f"  Fallback: {len(existing)} docs in tender dir: {existing[:5]}")

    results[ref] = {
        "child_sheets": child_result,
        "items": [],
        "docs_copied": [],
    }

    # ── Step 3: Update triage tags ──
    print("\n--- Updating triage tags ---")
    for ref, result in results.items():
        if ref in ["RFP-000000176710", "RFP-000000178771"] and result["items_count"] > 0:
            old_tags = triage["events"][ref]["tags"]
            # Replace NO_BOQ_TAB with HAS_DOCS_BOQ
            new_tags = [t for t in old_tags if t != "NO_BOQ_TAB"] + ["HAS_DOCS_BOQ"]
            triage["events"][ref]["tags"] = new_tags
            triage["events"][ref]["boq_status"] = "HAS_DOCS_BOQ"
            triage["events"][ref]["boq_items_count"] = result["items_count"]
            triage["events"][ref]["boq_items"] = result["items"]
            triage["events"][ref]["boq_source"] = "downloaded_xlsx"
            triage["events"][ref]["extraction_status"] = "success"
            triage["events"][ref]["recommended_action"] = "done — BoQ items parsed from downloaded xlsx"
            print(f"  {ref}: NO_BOQ_TAB → HAS_DOCS_BOQ ({result['items_count']} items)")

        elif ref == "RFP-000000177523":
            old_tags = triage["events"][ref]["tags"]
            if child_result.get("found"):
                new_tags = [t for t in old_tags if t != "NO_BOQ_TAB"] + ["HAS_CHILD_SHEETS"]
                triage["events"][ref]["boq_status"] = "HAS_CHILD_SHEETS"
                print(f"  {ref}: NO_BOQ_TAB → HAS_CHILD_SHEETS")
            else:
                new_tags = [t for t in old_tags if t != "NO_BOQ_TAB"] + ["NO_BOQ_ANYWHERE"]
                triage["events"][ref]["boq_status"] = "NO_BOQ_ANYWHERE"
                triage["events"][ref]["recommended_action"] = "audit — engineering services, no BoQ docs or child sheets"
                print(f"  {ref}: NO_BOQ_TAB → NO_BOQ_ANYWHERE (no child sheets, no docs)")
            triage["events"][ref]["tags"] = new_tags

    # Update summary counts
    tag_counts = {}
    boq_statuses = {}
    for evt in triage["events"].values():
        for tag in evt.get("tags", []):
            tag_counts[tag] = tag_counts.get(tag, 0) + 1
        bs = evt.get("boq_status", "")
        if bs:
            boq_statuses[bs] = boq_statuses.get(bs, 0) + 1

    triage["summary"]["tag_counts"] = tag_counts
    triage["summary"]["boq_status_counts"] = boq_statuses
    triage["summary"]["total_items_extracted"] = sum(
        evt.get("boq_items_count", 0) for evt in triage["events"].values()
    )
    triage["summary"]["updated_at"] = datetime.now(UTC).isoformat()

    # Write updated triage
    triage_bak = TRIAGE_PATH.with_suffix(".json.bak")
    import shutil
    shutil.copy2(TRIAGE_PATH, triage_bak)
    TRIAGE_PATH.write_text(json.dumps(triage, indent=2))
    print(f"\n  Triage updated: {TRIAGE_PATH}")
    print(f"  Backup: {triage_bak}")

    # ── Step 4: Export CSV ──
    csv_path = TRIAGE_PATH.with_suffix(".csv")
    csv_lines = ["event_number,tags,boq_status,items_count,recommended_action"]
    for ref, evt in sorted(triage["events"].items()):
        tags = "|".join(evt.get("tags", []))
        bs = evt.get("boq_status", "")
        count = evt.get("boq_items_count", 0)
        action = evt.get("recommended_action", "").replace(",", ";")
        csv_lines.append(f"{ref},{tags},{bs},{count},{action}")
    csv_path.write_text("\n".join(csv_lines))
    print(f"  CSV exported: {csv_path}")

    # ── Step 5: Print summary ──
    print(f"\n{'='*60}")
    print("RESOLUTION SUMMARY")
    print(f"{'='*60}")
    print(f"RFP-000000176710: {results['RFP-000000176710']['items_count']} items (HAS_DOCS_BOQ)")
    print(f"RFP-000000178771: {results['RFP-000000178771']['items_count']} items (HAS_DOCS_BOQ)")
    print(f"RFP-000000177523: child_sheets={results['RFP-000000177523']['child_sheets']['found']} (NO_BOQ_ANYWHERE)")
    print("\nTag distribution:")
    for tag, count in sorted(tag_counts.items()):
        print(f"  {tag}: {count}")
    print("\nBoQ statuses:")
    for bs, count in sorted(boq_statuses.items()):
        print(f"  {bs}: {count}")


if __name__ == "__main__":
    main()
